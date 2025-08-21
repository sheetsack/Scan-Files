# -*- coding: utf-8 -*-
"""
Сканер: первые N строк из файлов (csv/xls/xlsx) по папкам с авто-детектом кодировки CSV.

Функциональность:
- Настройки: Settings_Scan_Files.xlsx (лист "Settings"), рядом с .py/.exe.
  * Столбцы: Path, MaxFilesPerFolder (по умолчанию 100), LinesPerFile (по умолчанию 5).
  * Если файла нет — создаётся с дефолтами и скрипт завершает работу.
- Поддерживаемые расширения: .csv, .xls, .xlsx (регистронезависимо).
- В каждой папке/подпапке берём не более MaxFilesPerFolder файлов,
  сортируем по дате изменения (новые сверху), затем читаем первые N строк.
- CSV-файлы читаются с авто-детектом кодировки (BOM → charset-normalizer → chardet → fallback).
- Итог: CSV (ANSI cp1251), разделитель ';'.
  Колонки: File Path; Line 1; ...; Line N (N — максимум среди путей из настроек).
- Прогресс-бар через tqdm.

Зависимости:
  pip install openpyxl tqdm xlrd charset-normalizer chardet
"""

import os
import sys
import csv
import traceback
from datetime import datetime
from pathlib import Path
from typing import List, Tuple

# ---------- Константы и параметры ----------

SETTINGS_FILE = "Settings_Scan_Files.xlsx"
SETTINGS_SHEET = "Settings"

DEFAULT_MAX_FILES_PER_FOLDER = 100
DEFAULT_LINES_PER_FILE = 5

ALLOWED_EXTS = {".csv", ".xls", ".xlsx"}

# Сортировка файлов перед отсечением лимита
SORT_BY = "modified"   # 'modified' | 'created' | 'accessed'
SORT_DESC = True       # True — новые сверху

# ---------- tqdm (прогресс-бар) ----------

try:
    from tqdm import tqdm
except Exception:
    tqdm = None

# ---------- Вспомогательные функции ----------

def get_app_dir() -> Path:
    """Папка, где находится .py/.exe."""
    if getattr(sys, "frozen", False):  # PyInstaller
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def require_openpyxl() -> bool:
    try:
        import openpyxl  # noqa: F401
        return True
    except ImportError:
        print("Требуется 'openpyxl'. Установите: pip install openpyxl", file=sys.stderr)
        return False


def try_import_xlrd() -> bool:
    """Проверка наличия xlrd для .xls."""
    try:
        import xlrd  # noqa: F401
        return True
    except Exception:
        return False


def create_settings_file(path_to_file: Path):
    """Создать настройки с дефолтами и завершить работу."""
    if not require_openpyxl():
        sys.exit(1)
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = SETTINGS_SHEET

    ws["A1"] = "Path"
    ws["B1"] = "MaxFilesPerFolder"
    ws["C1"] = "LinesPerFile"
    ws["D1"] = "Комментарий"

    ws["A2"] = ""  # укажете путь
    ws["B2"] = DEFAULT_MAX_FILES_PER_FOLDER
    ws["C2"] = DEFAULT_LINES_PER_FILE
    ws["D2"] = "Заполните Path. Можно несколько строк (A2, A3, ...)."

    ws["D3"] = "MaxFilesPerFolder — лимит файлов на КАЖДУЮ папку/подпапку."
    ws["D4"] = "LinesPerFile — сколько первых строк брать из каждого файла."
    ws["D5"] = "Поддерживаемые расширения: csv, xls, xlsx."

    ws.column_dimensions["A"].width = 80
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 80

    wb.save(path_to_file)
    print(f"Создан файл настроек: {path_to_file}")
    print("Заполните лист 'Settings' и запустите скрипт снова.")
    sys.exit(0)


def read_settings(app_dir: Path) -> Tuple[List[Tuple[str, int, int]], int]:
    """
    Считать настройки.
    Возвращает:
      - список корней: [(path_str, max_files_per_folder, lines_per_file), ...]
      - глобальный максимум lines_per_file (для формирования заголовка CSV)
    Если путей нет — используем app_dir с дефолтами.
    """
    settings_path = app_dir / SETTINGS_FILE
    if not settings_path.exists():
        create_settings_file(settings_path)

    if not require_openpyxl():
        sys.exit(1)

    from openpyxl import load_workbook
    try:
        wb = load_workbook(settings_path, data_only=True)
    except Exception as e:
        print(f"Не удалось открыть {settings_path}: {e}", file=sys.stderr)
        sys.exit(1)

    if SETTINGS_SHEET not in wb.sheetnames:
        print(f"В {SETTINGS_FILE} нет листа '{SETTINGS_SHEET}'. Создаю и выхожу...")
        create_settings_file(settings_path)

    ws = wb[SETTINGS_SHEET]

    roots = []
    max_lines_global = 0

    for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
        path_cell = row[0]
        limit_cell = row[1]
        lines_cell = row[2]

        if path_cell is None or str(path_cell).strip() == "":
            continue

        p = os.path.expandvars(os.path.expanduser(str(path_cell).strip()))
        if not os.path.isdir(p):
            print(f"Предупреждение: путь не найден или не папка — {p}", file=sys.stderr)
            continue

        # Лимит
        limit = DEFAULT_MAX_FILES_PER_FOLDER
        if limit_cell is not None and str(limit_cell).strip() != "":
            try:
                v = int(float(str(limit_cell).strip()))
                if v > 0:
                    limit = v
            except Exception:
                print(f"Не удалось прочитать MaxFilesPerFolder '{limit_cell}' для {p}. Использую {DEFAULT_MAX_FILES_PER_FOLDER}.",
                      file=sys.stderr)

        # Линии
        lines = DEFAULT_LINES_PER_FILE
        if lines_cell is not None and str(lines_cell).strip() != "":
            try:
                v = int(float(str(lines_cell).strip()))
                if v > 0:
                    lines = v
            except Exception:
                print(f"Не удалось прочитать LinesPerFile '{lines_cell}' для {p}. Использую {DEFAULT_LINES_PER_FILE}.",
                      file=sys.stderr)

        roots.append((p, limit, lines))
        if lines > max_lines_global:
            max_lines_global = lines

    if not roots:
        print("В настройках нет корректных путей. Будет просканирована папка приложения.")
        roots = [(str(app_dir), DEFAULT_MAX_FILES_PER_FOLDER, DEFAULT_LINES_PER_FILE)]
        max_lines_global = DEFAULT_LINES_PER_FILE

    return roots, max_lines_global


def get_sort_key(stats: os.stat_result) -> float:
    if SORT_BY == "created":
        return float(stats.st_ctime)
    if SORT_BY == "accessed":
        return float(stats.st_atime)
    return float(stats.st_mtime)


# ---------- Автодетект кодировки CSV ----------

def detect_encoding(path: str) -> str | None:
    """
    Определить кодировку по первым байтам.
    Приоритет: BOM → charset-normalizer → chardet → None.
    Возвращает строку кодировки или None.
    """
    try:
        with open(path, "rb") as bf:
            data = bf.read(256 * 1024)  # до 256 КБ для анализа
    except Exception:
        return None

    # BOM
    if data.startswith(b"\xef\xbb\xbf"):
        return "utf-8-sig"
    if data.startswith(b"\xff\xfe\x00\x00") or data.startswith(b"\x00\x00\xfe\xff"):
        return "utf-32"  # Python сам учтёт порядок байт по BOM
    if data.startswith(b"\xff\xfe") or data.startswith(b"\xfe\xff"):
        return "utf-16"

    # charset-normalizer
    try:
        from charset_normalizer import from_bytes  # type: ignore
        result = from_bytes(data)
        if result is not None:
            best = result.best()
            if best and best.encoding:
                enc = best.encoding
                # ascii -> безопасно читать как utf-8
                if enc.lower() == "ascii":
                    return "utf-8"
                return enc
    except Exception:
        pass

    # chardet
    try:
        import chardet  # type: ignore
        guess = chardet.detect(data)
        enc = guess.get("encoding")
        if enc:
            if enc.lower() == "ascii":
                return "utf-8"
            return enc
    except Exception:
        pass

    return None


def read_csv_first_lines(path: str, n: int) -> List[str]:
    """
    Прочитать первые n строк CSV как обычный текст с авто-детектом кодировки.
    Возвращает список строк без переводов строк.
    """
    enc = detect_encoding(path)
    # Порядок попыток: определённая → типичные
    candidates = []
    if enc:
        candidates.append(enc)
    candidates += ["utf-8-sig", "cp1251", "utf-16", "latin1"]

    for encoding in candidates:
        try:
            lines = []
            with open(path, "r", encoding=encoding, errors="strict", newline="") as f:
                for i, line in enumerate(f):
                    if i >= n:
                        break
                    lines.append(line.rstrip("\r\n"))
            return lines
        except Exception:
            continue

    # Последняя попытка — максимально лояльно
    try:
        lines = []
        with open(path, "r", encoding="cp1251", errors="replace", newline="") as f:
            for i, line in enumerate(f):
                if i >= n:
                    break
                lines.append(line.rstrip("\r\n"))
        return lines
    except Exception:
        return ["<не удалось прочитать CSV>"]


def read_xlsx_first_rows(path: str, n: int) -> List[str]:
    """Первые n строк из .xlsx (первый лист)."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb.worksheets[0]
        out = []
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=n, values_only=True), start=1):
            values = [] if row is None else ["" if v is None else str(v) for v in row]
            out.append("\t".join(values).rstrip())
        if not out:
            out = ["" for _ in range(n)]
        return out
    except Exception as e:
        return [f"<xlsx ошибка: {e}>"]


def read_xls_first_rows(path: str, n: int) -> List[str]:
    """Первые n строк из .xls (первый лист). Требуется xlrd."""
    if not try_import_xlrd():
        return ["<xlrd не установлен. Установите: pip install xlrd>"]
    try:
        import xlrd
        book = xlrd.open_workbook(path)
        sheet = book.sheet_by_index(0)
        out = []
        rows = min(n, sheet.nrows)
        for r in range(rows):
            row_vals = sheet.row_values(r)
            out.append("\t".join("" if v is None else str(v) for v in row_vals).rstrip())
        if rows == 0:
            out = ["" for _ in range(n)]
        return out
    except Exception as e:
        return [f"<xls ошибка: {e}>"]


def get_first_lines_for_file(path: str, n: int) -> List[str]:
    ext = Path(path).suffix.lower()
    if ext == ".csv":
        return read_csv_first_lines(path, n)
    elif ext == ".xlsx":
        return read_xlsx_first_rows(path, n)
    elif ext == ".xls":
        return read_xls_first_rows(path, n)
    return ["<неподдерживаемое расширение>"]


def count_expected_files(roots_cfg: List[Tuple[str, int, int]]) -> int:
    """Подсчитать ожидаемое количество файлов (с учётом расширений и лимита на папку)."""
    total = 0
    for base, per_folder_limit, _lines in roots_cfg:
        for root, _dirs, files in os.walk(base, topdown=True, onerror=None, followlinks=False):
            candidates = [f for f in files if Path(f).suffix.lower() in ALLOWED_EXTS]
            # сортировка по дате
            try:
                items = []
                for name in candidates:
                    fp = os.path.join(root, name)
                    try:
                        st = os.stat(fp)
                        items.append((name, get_sort_key(st)))
                    except Exception:
                        pass
                items.sort(key=lambda t: t[1], reverse=SORT_DESC)
                count_here = len(items) if per_folder_limit <= 0 else min(per_folder_limit, len(items))
            except Exception:
                count_here = len(candidates) if per_folder_limit <= 0 else min(per_folder_limit, len(candidates))
            total += count_here
    return total


def scan_and_dump(roots_cfg: List[Tuple[str, int, int]], out_dir: Path, global_lines_max: int) -> Path:
    """
    Сканируем и записываем CSV.
    Первая колонка: полный путь к файлу с именем и расширением.
    Далее: Line 1..Line N (N = global_lines_max).
    """
    out_dir.mkdir(parents=True, exist_ok=True)
    out_csv = out_dir / f"Files_Head_Scan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

    headers = ["File Path"] + [f"Line {i}" for i in range(1, global_lines_max + 1)]

    total_written = 0
    skipped = 0

    bar = None
    if tqdm is not None:
        try:
            expected = count_expected_files(roots_cfg)
            bar = tqdm(total=expected, unit="file", ascii=True, dynamic_ncols=True, leave=False)
        except Exception:
            bar = None

    with open(out_csv, "w", encoding="cp1251", errors="replace", newline="") as f:
        writer = csv.writer(f, delimiter=";", quoting=csv.QUOTE_MINIMAL, lineterminator="\n")
        writer.writerow(headers)

        for base, per_folder_limit, lines_per_file in roots_cfg:
            for root, _dirs, files in os.walk(base, topdown=True, onerror=None, followlinks=False):
                # отбираем по расширениям
                candidates = [name for name in files if Path(name).suffix.lower() in ALLOWED_EXTS]

                # собираем статы для сортировки
                items = []
                for name in candidates:
                    fp = os.path.join(root, name)
                    try:
                        st = os.stat(fp)
                    except Exception:
                        continue
                    items.append((name, get_sort_key(st)))

                # сортируем и режем
                try:
                    items.sort(key=lambda t: t[1], reverse=SORT_DESC)
                except Exception:
                    pass

                if per_folder_limit > 0 and len(items) > per_folder_limit:
                    items = items[:per_folder_limit]

                # обрабатываем файлы
                for name, _k in items:
                    fp = os.path.join(root, name)
                    try:
                        lines = get_first_lines_for_file(fp, lines_per_file)
                        # нормализуем длину до global_lines_max
                        if len(lines) < global_lines_max:
                            lines = lines + [""] * (global_lines_max - len(lines))
                        elif len(lines) > global_lines_max:
                            lines = lines[:global_lines_max]

                        file_path = fp
                        if os.name == "nt":
                            file_path = file_path.replace("/", "\\")
                        writer.writerow([file_path] + lines)
                        total_written += 1
                    except Exception:
                        skipped += 1
                    finally:
                        if bar is not None:
                            try:
                                bar.update(1)
                            except Exception:
                                pass

    if bar is not None:
        try:
            bar.close()
        except Exception:
            pass

    print(f"Готово. Обработано файлов: {total_written}. Пропущено: {skipped}.")
    print(f"CSV сохранён: {out_csv}")
    if tqdm is None:
        print("Подсказка: установите 'tqdm' для прогресса: pip install tqdm")
    return out_csv

# ---------- Точка входа ----------

def main():
    app_dir = get_app_dir()
    roots_cfg, global_lines_max = read_settings(app_dir)
    scan_and_dump(roots_cfg, app_dir, global_lines_max)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nОстановлено пользователем.")
    except Exception:
        print("Произошла непредвиденная ошибка:", file=sys.stderr)
        traceback.print_exc()
        sys.exit(1)

